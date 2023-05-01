import pandas as pd
import openpyxl

filename = '1_health_and_beauty.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb.get_sheet_by_name('Sheet1')
for e in range(1, len(list(ws.iter_rows(values_only=True)))):
    ws[f"C{e}"] = ws[f"C{e}"].hyperlink.target if ws[f"C{e}"] and ws[f"C{e}"].hyperlink else ""

wb.save(filename)
