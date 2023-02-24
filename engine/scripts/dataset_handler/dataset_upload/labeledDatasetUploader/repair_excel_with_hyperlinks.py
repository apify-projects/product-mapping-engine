import pandas as pd
import openpyxl

wb = openpyxl.load_workbook('data/annotated_data/initial_files/eXtra -_ amazon - Petr Macek.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
for e in range(2, 302):
    ws[f"C{e}"] = ws[f"C{e}"].hyperlink.target if ws[f"C{e}"] and ws[f"C{e}"].hyperlink else ""

wb.save('data/annotated_data/initial_files/eXtra -_ amazon - Petr Macek.xlsx')
data = pd.read_excel('data/annotated_data/initial_files/eXtra -_ amazon - Petr Macek.xlsx')
data.to_csv('data/annotated_data/initial_files/amazon.csv', index=False)
'''
dataframe = pd.read_excel('Hotovo - Lednice - Tereza Plasova.xlsx')
dataframe.to_csv('Hotovo - Lednice - Tereza Plasova.csv')
'''