import pandas as pd
import openpyxl

wb = openpyxl.load_workbook('Hotovo - Lednice - Tereza Plasova (1).xlsx')
ws = wb.get_sheet_by_name('Sheet1')
for e in range(2, 302):
    ws[f"C{e}"] = ws[f"C{e}"].hyperlink.target if ws[f"C{e}"] and ws[f"C{e}"].hyperlink else ""

wb.save('Hotovo - Lednice - Tereza Plasova.xlsx')
dataframe = pd.read_excel('Hotovo - Lednice - Tereza Plasova.xlsx')
dataframe.to_csv('Hotovo - Lednice - Tereza Plasova.csv')